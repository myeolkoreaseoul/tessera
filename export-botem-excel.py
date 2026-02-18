#!/usr/bin/env python3
"""
보탬e 판정 결과 → 엑셀 + 표준 의견 텍스트 생성
- results.json + data.json 병합
- 검토의견(opinion) 필드 추가 → results.json 업데이트
- 엑셀 파일: projects/캠퍼스타운-고려대/검토결과.xlsx
"""
import json
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR     = Path(__file__).parent / 'projects/캠퍼스타운-고려대'
DATA    = DIR / 'data.json'
RESULTS = DIR / 'results.json'
OUTPUT  = DIR / '검토결과.xlsx'

# ───────────────────────────────────────────────────────────────────
# 비목별 표준 의견 텍스트 (보탬e "보완요청" 입력용)
# ───────────────────────────────────────────────────────────────────
OPINION_TEMPLATE = {
    '강사비(2-1)':
        '보완요청: ① 이력서(최초1회) ② 강의확인서(서명) ③ 강의증빙(교안/사진/출석부) ④ 계좌이체확인증 ⑤ 원천징수영수증(125,000원 초과시)',
    '멘토비(2-3)':
        '보완요청: ① 멘토링내역서(서명) ② 이력서(최초1회) ③ 계좌이체확인증 ④ 1일단가(대면30만/비대면15만) 및 월200만원 한도 확인 ⑤ 원천징수영수증(125,000원 초과시)',
    '심사평가비(2-2)':
        '보완요청: ① 참석자명부(서명) ② 심사평가결과서 ③ 이력서(최초1회) ④ 계좌이체확인증 ⑤ 단가확인(시간당10만, 1일50만, 서면7만) ⑥ 전담조직소속 여부',
    '회의비/자문비(2-4)':
        '보완요청: ① 회의참석확인서(서명) ② 회의록 ③ 이력서(최초1회) ④ 계좌이체확인증 ⑤ 단가확인(대면15/20만, 비대면5/7만) ⑥ 전담조직소속 여부',
    '단순인건비(2-5)':
        '보완요청: ① 활동기록부(서명) ② 지급내역서 ③ 계좌이체확인증 ④ 근무시간 확인(8h/일, 60h/월) ⑤ 서울형 생활임금 단가',
    '프로그램용역비(2-6)':
        '보완요청: ① 프로그램기획서 ② 견적서(비교견적서) ③ 사업자등록증 ④ 검수조서 ⑤ 카드영수증/계좌이체확인증 ⑥ 업무범위 구분 ⑦ 2천만원 이상시 나라장터',
    '단기임차료(2-7)':
        '보완요청: ① 행사개요(임차목적/일시) ② 견적서(비교견적서) ③ 카드영수증/계좌이체확인증',
    '행사비(2-8)':
        '보완요청: ① 견적서(비교견적서) ② 행사사진 ③ 카드영수증/계좌이체확인증',
    '식비/다과비(2-9)':
        '보완요청: ① 활동내역서/회의록 ② 참석자명단(서명) ③ 카드영수증 ④ 1인단가 확인(식비8천/다과4천) ⑤ 외부인 참석여부',
    '창업지원금(3-1)':
        '보완요청: ① 수령확인서(서명) ② 지출증빙서류 ③ 창업활동기록부 ④ 공간활용계획서 ⑤ 계좌이체확인증 ⑥ 팀당 연3천만원 한도 ⑦ 인건비/자산취득비 해당여부',
    '창업시상금(3-2)':
        '보완요청: ① 추진계획서(시상계획) 및 결과보고서 ② 수령확인서(서명) ③ 계좌이체확인증 ④ 연1천만원/1팀3백만원 한도 ⑤ 경진대회 선정결과',
    '전담인력인건비(4-2)':
        '보완요청: ① 근로계약서/이력서(최초1회) ② 급여지급기준 증빙(최초1회) ③ 지급내역서 ④ 계좌이체확인증 ⑤ 원천징수영수증 ⑥ 기간제 여부(정규직 불가) ⑦ 1인6천만/시비15% 한도',
    '여비(5-4)':
        '보완요청: ① 활동결과보고서 ② 지급내역서(서명) ③ 계좌이체확인증 ④ 시내여비 단가(4h미만1만/4h이상2만) ⑤ 시외여비시 사전승인문서',
    '공간운영비(5-2)':
        '보완요청: ① 공과금 영수증(납부확인서) ② 카드영수증/세금계산서/계좌이체확인증 ③ 입주기업 창업활동 공간 확인',
    '공간운영비(시설수선비)(5-2)':
        '보완요청: ① 카드영수증/세금계산서 ② 인테리어 아님 확인 ③ 임차시설/전담조직전용시설 여부(해당시 수선비 불가) ④ 기본기능 유지보수 목적 확인',
    '소모성물품(5-3)':
        '보완요청: ① 구입영수증/내역서 ② 카드영수증/계좌이체확인증 ③ 공간상주인원 x 월5만원 한도 ④ 견적서(100만원 이상)',
    '홍보비(5-7)':
        '보완요청: ① 견적서(비교견적서) ② 홍보물 설치사진 ③ 인쇄물3부 ④ 카드영수증/세금계산서 ⑤ 책자형시 배포계획서',
    '물품임차비(5-8)':
        '보완요청: ① 임대차계약서 ② 견적서(비교견적서) ③ 카드영수증/세금계산서 ④ 세부실행계획서 승인확인 ⑤ 시비6천만원 한도',
    '시설비(6-1)':
        '보완요청: ① 견적서(비교견적서) ② 공사설계도면/내역서 ③ 공사전후사진 ④ 영수증/세금계산서 ⑤ 세부실행계획서 승인확인',
    '물품취득비(6-2)':
        '보완요청: ① 구입내역서 ② 견적서(비교견적서) ③ 물품관리대장 ④ 카드영수증/세금계산서 ⑤ 세부실행계획서 승인확인',
    '교육훈련비(4-3)':
        '보완요청: ① 교육참가신청서(서명) ② 교육이수증 ③ 계좌이체확인증 ④ 사업관련성 확인 ⑤ 연200만원 한도',
    '회계감사비(5-6)':
        '보완요청: ① 회계감사보고서 ② 500만원 한도 확인',
    '사무관리비(5-3/5-7)':
        '보완요청: 세부비목 확인 후 해당 증빙 보완 (소모품: 구입영수증/한도 / 홍보비: 견적서/설치사진 / 행사비: 견적서/사진)',
    '교육연구프로그램(2-10)':
        '보완요청: ① 프로그램기획서 ② 참석자서명부 ③ 카드영수증 ④ 창업인재육성 교과/비교과 프로그램 해당여부',
    '사업단장수당(4-1)':
        '보완요청: ① 지급내역서 ② 계좌이체확인증 ③ 대학내부규정 지급근거(최초1회) ④ 월100만원 한도 ⑤ 겸직수당 이중수령여부',
}

# ───────────────────────────────────────────────────────────────────
# 의견 텍스트 생성
# ───────────────────────────────────────────────────────────────────
def get_opinion(result):
    item_type = result.get('type', '')
    issues    = result.get('issues', [])
    status    = result.get('status', '확인')

    if status == 'SKIP':
        return '전자세금계산서 연동 증빙 확인'

    # 즉시 확인 사항이 있으면 앞에 추가
    alert_msgs = [i for i in issues if '⚠️' in i]

    # 비목별 표준 의견 찾기
    base_opinion = OPINION_TEMPLATE.get(item_type)

    # 키 부분 매칭 (비목 앞부분)
    if not base_opinion:
        for key, val in OPINION_TEMPLATE.items():
            if item_type.startswith(key.split('(')[0]):
                base_opinion = val
                break

    if not base_opinion:
        base_opinion = '보완요청: 증빙서류(계획서, 영수증, 계좌이체확인증 등) 보완 필요'

    # 즉시 확인 사항 앞에 붙임
    if alert_msgs:
        prefix = ' '.join(alert_msgs) + ' / '
        return prefix + base_opinion

    return base_opinion

# ───────────────────────────────────────────────────────────────────
# 엑셀 생성
# ───────────────────────────────────────────────────────────────────
STATUS_COLOR = {
    '적정': 'C6EFCE',
    '확인': 'FFEB9C',
    'SKIP': 'DDEBF7',
}

def make_excel(data, results):
    rn_to_item = {item['순번']: item for item in data}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '검토결과'

    # 헤더
    headers = [
        '순번', '집행일자', '비목', '집행목적(용도)', '거래처', '지방비금액', '자부담금액',
        '집행방식', '상태', '확인사항', '표준검토의견',
        '계약등록', '중요재산'
    ]
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # 데이터 행
    for r in results:
        rn  = str(r['rowNum'])
        row_data = rn_to_item.get(rn, {})
        opinion  = get_opinion(r)

        row = [
            r['rowNum'],
            r['date'],
            r['type'],
            r['purpose'],
            r['vendor'],
            r['amount'],
            row_data.get('자부담 집행금액', '0'),
            row_data.get('집행방식', ''),
            r['status'],
            '\n'.join(r['issues']),
            opinion,
            row_data.get('계약정보등록여부', 'N'),
            row_data.get('중요재산등록여부', 'N'),
        ]

        ws_row = ws.max_row + 1
        fill_color = STATUS_COLOR.get(r['status'], 'FFFFFF')
        fill = PatternFill('solid', fgColor=fill_color)

        for col, val in enumerate(row, 1):
            cell = ws.cell(ws_row, col, val)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if col == 6:  # 지방비금액
                cell.number_format = '#,##0'
            if col == 9:  # 상태 컬럼에 색상
                cell.fill = fill

    # 컬럼 폭 설정
    col_widths = [6, 12, 22, 45, 18, 14, 12, 14, 8, 50, 80, 8, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 행 높이 자동 (최소 18)
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 40

    # 헤더 고정
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'저장: {OUTPUT}')

# ───────────────────────────────────────────────────────────────────
# results.json 에 opinion 필드 추가
# ───────────────────────────────────────────────────────────────────
def update_results(results):
    for r in results:
        r['opinion'] = get_opinion(r)
    RESULTS.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'results.json opinion 필드 추가 완료')

def main():
    data    = json.loads(DATA.read_text(encoding='utf-8'))
    results = json.loads(RESULTS.read_text(encoding='utf-8'))

    print(f'데이터: {len(data)}건, 판정: {len(results)}건')

    update_results(results)
    make_excel(data, results)

    # 통계
    counter = {}
    for r in results:
        s = r['status']
        counter[s] = counter.get(s, 0) + 1
    print(f'상태: {counter}')
    kb = OUTPUT.stat().st_size // 1024
    print(f'엑셀 크기: {kb}KB')

if __name__ == '__main__':
    main()
